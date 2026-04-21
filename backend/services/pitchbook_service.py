"""
PitchBook Excel Parser
Parses PitchBook "All Columns" exports into the expanded Master Universe schema.

PitchBook files have:
  - Metadata rows 1-7 (title, download date, search criteria)
  - Header row 8 (157 columns)
  - Data from row 9 onwards

Detection: filename contains "pitchbook" (case-insensitive)
"""
import re
import logging
from typing import List, Dict, Optional
from io import BytesIO
import openpyxl

logger = logging.getLogger(__name__)


# ── PitchBook Column → Master Universe Field Mapping ─────────────────────────
# Keys = PitchBook header text (row 8), Values = our BQ field name
PITCHBOOK_FIELD_MAP = {
    # ── Identity & Overview ──
    "Companies":                        "name",
    "Description":                      "description",
    "Website":                          "website",
    "LinkedIn URL":                     "linkedin_url",
    "Year Founded":                     "year_founded",
    "Employees":                        "employees",
    "Keywords":                         "keywords",
    "Verticals":                        "verticals",
    "Business Status":                  "business_status",

    # ── Sector / Industry ──
    "Primary PitchBook Industry Sector":"sector",
    "Primary PitchBook Industry Group": "industry_group",
    "Primary PitchBook Industry Code":  "industry_code",
    "Emerging Spaces":                  "emerging_spaces",

    # ── Geography ──
    "HQ Location":                      "hq_location",
    "HQ City":                          "hq_city",
    "HQ Country/Territory/Region":      "hq_country",
    "HQ Global Region":                 "hq_global_region",

    # ── Ownership & Financing ──
    "Ownership Status":                 "ownership",
    "Company Financing Status":         "financing_status",
    "Total Raised":                     "total_raised_m",
    "Financing Status Note":            "financing_note",

    # ── Financials ──
    "Revenue":                          "revenue_m",
    "EBITDA":                           "estimated_ebitda",
    "Net Income":                       "net_income_m",
    "Enterprise Value":                 "enterprise_value_m",
    "Revenue Growth %":                 "revenue_growth_pct",

    # ── Valuation ──
    "Valuation Estimate":               "valuation_estimate_m",
    "Last Known Valuation":             "last_valuation_m",
    "Last Known Valuation Date":        "last_valuation_date",

    # ── Primary Contact (founder/CEO) ──
    "Primary Contact":                  "contact_name",
    "Primary Contact Title":            "contact_title",
    "Primary Contact Email":            "contact_email",
    "Primary Contact Phone":            "contact_phone",
    "HQ Email":                         "hq_email",
    "HQ Phone":                         "hq_phone",

    # ── Investors ──
    "Active Investors":                 "active_investors",
    "# Active Investors":               "num_active_investors",
    "Former Investors":                 "former_investors",

    # ── Last Financing Round ──
    "Last Financing Date":              "last_financing_date",
    "Last Financing Size":              "last_financing_size_m",
    "Last Financing Valuation":         "last_financing_valuation_m",
    "Last Financing Deal Type":         "last_financing_type",

    # ── First Financing Round ──
    "First Financing Date":             "first_financing_date",
    "First Financing Size":             "first_financing_size_m",

    # ── Growth & Web Signals ──
    "Growth Rate":                      "pitchbook_growth_rate",
    "Growth Rate Percentile":           "growth_rate_percentile",
    "SimilarWeb Unique Visitors":       "web_visitors",

    # ── PitchBook Scoring ──
    "Opportunity Score":                "opportunity_score",
    "Success Probability":              "success_probability",
    "M&A Probability":                  "ma_probability",
    "Predicted Exit Type":              "predicted_exit_type",

    # ── Patents ──
    "Total Patent Documents":           "total_patents",

    # ── Other ──
    "Competitors":                      "competitors",
    "Company Also Known As":            "also_known_as",
    "Company Legal Name":               "legal_name",
    "Registration Number":              "registration_number",
}


def _extract_hyperlink(cell_value, prefer_label: bool = False) -> str:
    """
    PitchBook exports website/LinkedIn as =HYPERLINK("url", "label").

    prefer_label=True  → use the display text (e.g. "www.company.com"), add http:// if missing
    prefer_label=False → use the URL from the formula directly

    Falls back to the raw value if no HYPERLINK formula is found.
    """
    if cell_value is None:
        return ""
    val = str(cell_value).strip()

    # Match =HYPERLINK("url", "label") or =HYPERLINK("url")
    match = re.match(r'=HYPERLINK\("([^"]+)"(?:\s*,\s*"([^"]*)")?\)', val)
    if match:
        url = match.group(1)
        label = match.group(2) or ""

        if prefer_label and label:
            # Use the display label — it's the company's actual domain
            result = label.strip()
            if result and not result.startswith("http"):
                result = "https://" + result
            return result
        return url

    # No formula — return as-is, ensure URL format
    if val and not val.startswith("http") and ("." in val):
        val = "https://" + val
    return val


def _safe_float(val, default=0.0) -> float:
    if val is None or val == "" or str(val).strip() == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=None) -> Optional[int]:
    if val is None or val == "" or str(val).strip() == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _map_pitchbook_region(hq_country: str, global_region: str) -> str:
    """Map PitchBook geography fields to our standard region values."""
    country = (hq_country or "").lower()
    region = (global_region or "").lower()

    if "united kingdom" in country or "uk" in country:
        return "UK"
    if "ireland" in country:
        return "Ireland"
    if "europe" in region:
        return "Europe"
    if "north america" in region:
        return "North America"
    if region:
        return region.title()
    return "Unknown"


def _map_pitchbook_ownership(ownership_status: str, financing_status: str) -> str:
    """Map PitchBook ownership to our standard values."""
    own = (ownership_status or "").lower()
    fin = (financing_status or "").lower()

    if "no backing" in own:
        return "Bootstrapped"
    if "angel" in fin:
        return "Angel-backed"
    if "venture" in fin or "vc" in fin:
        return "VC-backed"
    if "pe" in fin or "buyout" in fin:
        return "PE-backed"
    if "backing" in own:
        return "Backed"
    return _safe_str(ownership_status) or "Unknown"


def parse_pitchbook_excel(file_bytes: bytes) -> List[Dict]:
    """
    Parse a PitchBook 'All Columns' Excel export.

    Returns a list of dicts matching the expanded Master Universe schema.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb.active

    # ── Step 1: Find the header row (usually row 8) ──
    header_row_idx = None
    for row_idx in range(1, 15):
        cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        values = [c.value for c in cells if c.value is not None]
        # The header row has "Companies" and "Description" among many non-null values
        str_values = [str(v) for v in values]
        if "Companies" in str_values and "Description" in str_values:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        logger.warning("Could not find PitchBook header row. Falling back to row 8.")
        header_row_idx = 8

    # ── Step 2: Build column index ──
    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
    headers = [c.value for c in header_cells]

    # Map: column_index → our_field_name
    col_to_field = {}
    for col_idx, header_val in enumerate(headers):
        if header_val and str(header_val) in PITCHBOOK_FIELD_MAP:
            col_to_field[col_idx] = PITCHBOOK_FIELD_MAP[str(header_val)]

    logger.info(f"PitchBook parser: header at row {header_row_idx}, mapped {len(col_to_field)}/{len(headers)} columns")

    # ── Step 3: Parse data rows ──
    companies = []
    data_start = header_row_idx + 1

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        raw = {}
        for col_idx, field_name in col_to_field.items():
            if col_idx < len(row):
                raw[field_name] = row[col_idx].value

        name = _safe_str(raw.get("name"))
        if not name:
            continue

        # Extract hyperlinks — use display label for website (it's the actual domain)
        # Use URL for LinkedIn (the full linkedin.com/company/... path)
        website = _extract_hyperlink(raw.get("website"), prefer_label=True)
        linkedin = _extract_hyperlink(raw.get("linkedin_url"), prefer_label=False)

        # Use actual country name as region (not short code)
        hq_country = _safe_str(raw.get("hq_country"))
        global_region = _safe_str(raw.get("hq_global_region"))
        region = hq_country if hq_country else global_region

        ownership_raw = _safe_str(raw.get("ownership"))
        financing_raw = _safe_str(raw.get("financing_status"))
        ownership = _map_pitchbook_ownership(ownership_raw, financing_raw)

        company = {
            # ── Existing fields ──
            "name": name,
            "website": website,
            "sector": _safe_str(raw.get("sector")),
            "region": region,
            "ownership": ownership,
            "description": _safe_str(raw.get("description")),
            "match_score": 0.0,
            "status": "Uploaded",
            "contact_name": _safe_str(raw.get("contact_name")),
            "contact_email": _safe_str(raw.get("contact_email")),
            "linkedin_url": linkedin,
            "growth_signals": bool(raw.get("pitchbook_growth_rate")),
            "estimated_ebitda": _safe_float(raw.get("estimated_ebitda")),

            # ── New expanded fields ──
            "contact_title": _safe_str(raw.get("contact_title")),
            "contact_phone": _safe_str(raw.get("contact_phone")),
            "hq_email": _safe_str(raw.get("hq_email")),
            "hq_phone": _safe_str(raw.get("hq_phone")),
            "hq_location": _safe_str(raw.get("hq_location")),
            "hq_city": _safe_str(raw.get("hq_city")),
            "hq_country": hq_country,
            "employees": _safe_int(raw.get("employees")),
            "year_founded": _safe_int(raw.get("year_founded")),
            "keywords": _safe_str(raw.get("keywords")),
            "verticals": _safe_str(raw.get("verticals")),
            "industry_group": _safe_str(raw.get("industry_group")),
            "industry_code": _safe_str(raw.get("industry_code")),
            "emerging_spaces": _safe_str(raw.get("emerging_spaces")),
            "business_status": _safe_str(raw.get("business_status")),
            "financing_status": financing_raw,
            "total_raised_m": _safe_float(raw.get("total_raised_m")),
            "revenue_m": _safe_float(raw.get("revenue_m")),
            "net_income_m": _safe_float(raw.get("net_income_m")),
            "enterprise_value_m": _safe_float(raw.get("enterprise_value_m")),
            "revenue_growth_pct": _safe_float(raw.get("revenue_growth_pct")),
            "valuation_estimate_m": _safe_float(raw.get("valuation_estimate_m")),
            "last_valuation_m": _safe_float(raw.get("last_valuation_m")),
            "last_valuation_date": _safe_str(raw.get("last_valuation_date")),
            "active_investors": _safe_str(raw.get("active_investors")),
            "num_active_investors": _safe_int(raw.get("num_active_investors")),
            "former_investors": _safe_str(raw.get("former_investors")),
            "last_financing_date": _safe_str(raw.get("last_financing_date")),
            "last_financing_size_m": _safe_float(raw.get("last_financing_size_m")),
            "last_financing_valuation_m": _safe_float(raw.get("last_financing_valuation_m")),
            "last_financing_type": _safe_str(raw.get("last_financing_type")),
            "first_financing_date": _safe_str(raw.get("first_financing_date")),
            "first_financing_size_m": _safe_float(raw.get("first_financing_size_m")),
            "pitchbook_growth_rate": _safe_float(raw.get("pitchbook_growth_rate")),
            "growth_rate_percentile": _safe_int(raw.get("growth_rate_percentile")),
            "web_visitors": _safe_int(raw.get("web_visitors")),
            "opportunity_score": _safe_int(raw.get("opportunity_score")),
            "success_probability": _safe_int(raw.get("success_probability")),
            "ma_probability": _safe_int(raw.get("ma_probability")),
            "predicted_exit_type": _safe_str(raw.get("predicted_exit_type")),
            "total_patents": _safe_int(raw.get("total_patents")),
            "competitors": _safe_str(raw.get("competitors")),
            "also_known_as": _safe_str(raw.get("also_known_as")),
            "legal_name": _safe_str(raw.get("legal_name")),
            "registration_number": _safe_str(raw.get("registration_number")),
            "financing_note": _safe_str(raw.get("financing_note")),
        }

        companies.append(company)

    wb.close()
    logger.info(f"PitchBook parser: extracted {len(companies)} companies")
    return companies
